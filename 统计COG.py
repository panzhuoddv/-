import openpyxl
from collections import Counter
import os
path=r'd:\Study\研究生\文章\NBD domain drives the functional specificity of 3192 and 6671\20210807\Figure\Figure 2\汇总统计规律.xlsx'
# files= os.listdir(path)
# for file in files:
#     wb = openpyxl.load_workbook(path+'/'+file)
#     sh=wb['Sheet1']
#     a=[]
#     for j in range(10,14):
#         for i in range(2,sh.max_row+1):
#             a.append(sh.cell(i,j).value)
#     dict={}        
#     for key in a:
#         dict[key] = dict.get(key, 0) + 1
#     print(dict)
#     n=2
#     m=10
#     g=wb['Sheet3']
#     g.cell(n-1,m-1).value='COG'
    
#     for i in dict.keys():
#         g.cell(n,m-1).value=i
#         g.cell(n,m).value=dict[i]
#         n+=1
#     wb.save(path+'/'+file)
wb = openpyxl.load_workbook(path)
A = '3192&6671'

sh=wb[A]
a=[]

for j in range(11,15):
    for i in range(2,sh.max_row+1):
        a.append(sh.cell(i,j).value)
dict={}        
for key in a:
    dict[key] = dict.get(key, 0) + 1
print(dict)
n=2
m=19
g=wb[A]
g.cell(n-1,m-1).value='COG'

for i in dict.keys():
    g.cell(n,m-1).value=i
    g.cell(n,m).value=dict[i]
    n+=1
wb.save(path)