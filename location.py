import openpyxl
wb = openpyxl.load_workbook(r'd:\Study\研究生\实验\CO-IP\正常株\正常培养条件\6671\E19.xlsx')
sh = wb['减去阴性对照']
# sh.cell(1,20).value = 'sublocation'
m = sh.max_row
ids=[]
for i in range(1,m):
    ids.append(sh.cell(i,9).value)
h=1

import requests # 调用requests库
import re,random
import time
from random import choice
from bs4 import BeautifulSoup # 调用BeautifulSoup库
file = open(r'D:\Study\python\manucript\random requests head.txt', 'r')
user_agent_list = file.readlines()
for id in ids:
    h+=1
    user_agent = str(choice(user_agent_list)).replace('\n', '')
    user_agent = 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:39.0) Gecko/20100101 Firefox/39.0' if len(
        user_agent) < 10 else user_agent
    headers = {
        'Host': "map.baidu.com",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "en-US,en;q=0.5",
        "Connection": "keep-alive",
        "User-Agent": user_agent
    }
    try:
        res =requests.get('http://db.psort.org/search/results?show_adjust=1&dataset=c&refseq='+id+'&protein=&organism=&phylum=&class=&order=&family=&genus=',headers=headers,timeout=500)
    except requests.exceptions.ConnectionError:
        sh.cell(h,11).value='error'
        print('error')
        continue
    
    bs = BeautifulSoup(res.text,'html.parser')
    location=bs.find_all('p', class_='card-text')
    
    sh.cell(h,11).value=re.sub('[PSORTb v 3.00]','',location[16].text)
    print('percent: {:.2%}'.format(h/m))
    print(h)
    if h//50 ==1 :
        wb.save(r'd:\Study\研究生\实验\CO-IP\正常株\正常培养条件\3192\E3.xlsx')
wb.save(r'd:\Study\研究生\实验\CO-IP\正常株\正常培养条件\3192\E3.xlsx')
