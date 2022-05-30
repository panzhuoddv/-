import openpyxl
from collections import Counter
import os
path=r'd:\Study\研究生\实验\CO-IP\scope新旧蛋白号23.xlsx'
# files= os.listdir(path)
# for file in files:
wb = openpyxl.load_workbook(path)
sh=wb['scope']
a=[]
for j in range(15,26):
    for i in range(2,sh.max_row+1):
        a.append(sh.cell(i,j).value)
dict={}        
for key in a:
    dict[key] = dict.get(key, 0) + 1
print(dict)
n=2
m=13
g=wb['Sheet2']
g.cell(n-1,m-1).value='Scope fold'
for i in dict.keys():
    g.cell(n,m-1).value=i
    g.cell(n,m).value=dict[i]
    n+=1
wb.save(path)