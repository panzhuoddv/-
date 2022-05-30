import openpyxl
wb = openpyxl.load_workbook(r'd:\Study\研究生\实验\CO-IP\DK1622蛋白location+COG功能注释.xlsx')
sh = wb['Sheet1']
sh.cell(1,13).value = 'COG_class'
m = sh.max_row
ids=[]
for i in range(2,m+1):
    ids.append(sh.cell(i,9).value)
# print(ids)

import requests # 调用requests库
from bs4 import BeautifulSoup # 调用BeautifulSoup库
from fake_useragent import UserAgent
import random
user_agent = UserAgent()
cog=[]
h=1
for id in ids:
    user_agent = UserAgent()
    headers = {
        "User-Agent": user_agent.random
    }
    
    res =requests.get('https://www.ncbi.nlm.nih.gov/research/cog/search/?query=prot%3A'+id,headers=headers)
    # print(res.status_code)
    bs = BeautifulSoup(res.text,'html.parser')
    list2=bs.find('tbody')
    
    # print(list2[3].text)
    # print(list)
    h=h+1
    print(h)
    # print(id)
    n=-1
    # print(list)
    if not list2 is None:
        list=list2.find_all('td')
        if not list is None:
            m=0
            for i in list[3].text:
                n=n+1
                if list[3].text[n-1]==' ' and list[3].text[n]!=' ' and list[3].text[n]!='\n' and list[3].text[n+2]==' ':
                    if list[3].text[n]=='J'or list[3].text[n]=='A'or list[3].text[n]=='K'or list[3].text[n]=='L' or list[3].text[n]=='B'or list[3].text[n]=='D'or list[3].text[n]=='Y'or list[3].text[n]=='V'or list[3].text[n]=='T'or list[3].text[n]=='M'or list[3].text[n]=='N'or list[3].text[n]=='Z'or list[3].text[n]=='W'or list[3].text[n]=='U'or list[3].text[n]=='O'or list[3].text[n]=='X'or list[3].text[n]=='C'or list[3].text[n]=='G'or list[3].text[n]=='E'or list[3].text[n]=='F'or list[3].text[n]=='H'or list[3].text[n]=='I'or list[3].text[n]=='P'or list[3].text[n]=='Q'or list[3].text[n]=='R'or list[3].text[n]=='S':
                        sh.cell(h,13+m).value=list[3].text[n]
                        # print(list[3].text[n])
                        m=m+1
                        # print(list.text[n])
                    else:
                        continue
                    # print(list.text[n])
        else:
            sh.cell(h,13).value='None'
    else:
        sh.cell(h,13).value='None'
    if h % 50==1:
        wb.save(r'd:\Study\研究生\实验\CO-IP\DK1622蛋白location+COG功能注释.xlsx')
        wb = openpyxl.load_workbook(r'd:\Study\研究生\实验\CO-IP\DK1622蛋白location+COG功能注释.xlsx')
        sh = wb['Sheet1']
            
wb.save(r'd:\Study\研究生\实验\CO-IP\DK1622蛋白location+COG功能注释.xlsx')