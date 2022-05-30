import openpyxl,re
path=r'C:\Users\hp\Desktop\新建 Microsoft Excel 工作表.xlsx'
wb = openpyxl.load_workbook(path)
sh = wb['Sheet1']
sh.cell(1,3).value = 'Gravy'
m = sh.max_row
ids=[]
for i in range(2,m+1):
    ids.append(sh.cell(i,2).value)
# print(ids)

import requests # 调用requests库
from bs4 import BeautifulSoup # 调用BeautifulSoup库

import random
# user_agent = UserAgent()
cog=[]
h=4252
for id in ids[4251:]:
    h+=1
    # user_agent = UserAgent()
    # headers = {
    #     "User-Agent": user_agent.random
    # }
    res =requests.get('https://web.expasy.org/cgi-bin/protparam/protparam1?'+id+'@noft@')
    bs = BeautifulSoup(res.text,'html.parser')
    list=bs.find('pre')
    A=list.text
    # print(A[A.find('(GRAVY): ')+9:])
    sh.cell(h,10).value=re.sub('\n','',A[A.find('(GRAVY): ')+9:])
    print('percent: {:.2%}'.format(h/m)+'----------'+re.sub('\n','',A[A.find('(GRAVY): ')+9:]))
    if h % 50==1:
        wb.save(path)
        wb = openpyxl.load_workbook(path)
        sh = wb['Sheet1']
            
wb.save(path)
# print(a)
# print(' ')
# print('o')
# print(' ')