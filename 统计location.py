import openpyxl
from collections import Counter
import os
path=r'D:\Study\研究生\实验\CO-IP\正常株\20201123\压力条件'
files= os.listdir(path)
for file in files:

    wb = openpyxl.load_workbook(path+'/'+ file)
    sh=wb['Sheet1']
    a=[]
    m=5
    for i in range(2,sh.max_row+1):
        a.append(sh.cell(i,m).value)

    dict={}        
    for key in a:
        dict[key] = dict.get(key, 0) + 1
    print(dict)
    n=2
    g=wb['Sheet3']
    g.cell(n-1,m-1).value='heat'
    for i in dict.keys():
        g.cell(n,m-1).value=i
        g.cell(n,m).value=dict[i]
        n+=1
    wb.save(path+'/'+ file)